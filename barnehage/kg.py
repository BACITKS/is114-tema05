from flask import Flask
from flask import url_for
from flask import render_template
from flask import request
from flask import redirect
from flask import session
from kgmodel import (Foresatt, Barn, Soknad, Barnehage)
from kgcontroller import (form_to_object_soknad, insert_soknad, commit_all, select_alle_barnehager)
import altair as alt
import pandas as pd
from io import StringIO

app = Flask(__name__)
app.secret_key = 'BAD_SECRET_KEY' # nødvendig for session

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/barnehager')
def barnehager():
    information = select_alle_barnehager()
    return render_template('barnehager.html', data=information)

@app.route('/behandle', methods=['GET', 'POST'])
def behandle():
    if request.method == 'POST':
        sd = request.form
        print(sd)
        log = insert_soknad(form_to_object_soknad(sd))
        print(log)
        
        # Legg til logikk
        antall_ledige_plasser = 5 
        fortrinnsrett_grunner = [
            sd.get('fortrinnsrett_barnevern') == 'on',
            sd.get('fortrinnsrett_sykdom_familien') == 'on',
            sd.get('fortrinnsrett_sykdom_barnet') == 'on',
            sd.get('fortrinnsrett_annet') == 'on'
        ]
        har_fortrinnsrett = any(fortrinnsrett_grunner)

        # Bestem resultatet basert på ledige plasser og fortrinnsrett
        if antall_ledige_plasser > 0 or har_fortrinnsrett:
            resultat = "TILBUD"
        else:
            resultat = "AVSLAG"

        session['information'] = sd
        session['resultat'] = resultat  # Lagre resultatet i session
        return redirect(url_for('svar'))
    else:
        return render_template('soknad.html')

import openpyxl

def hent_alle_soeknader():
    workbook = openpyxl.load_workbook(r"C:\oblig5\is114-tema05\barnehage\kgdata.xlsx")
    sheet = workbook.active
    soeknader = []

    # Hardkodet liste med barnehageinformasjon for å matche `barnehage_id` til `barnehage_navn`
    barnehager = {
        1: 'Sunshine Preschool',
        2: 'Happy Days Nursery',
        3: '123 Learning Center',
        4: 'ABC Kindergarten',
        5: 'Tiny Tots Academy',
        6: 'Giggles and Grins Childcare',
        7: 'Playful Pals Daycare'
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        barnehage_id = row[0]  # Første kolonne er barnehage_id
        soeknad = {
            'barnehage_id': barnehage_id,
            'navn_foresatt': row[1],     # Navn på foresatt
            'adresse': row[2],           # Adresse
            'telefon': row[3],           # Telefon
            'barnehage_navn': barnehager.get(barnehage_id, "Ukjent")  # Hent barnehagenavn fra `barnehager`
        }
        soeknader.append(soeknad)

    return soeknader

def hent_alle_barnehager():
    workbook = openpyxl.load_workbook('C:\\oblig5\\is114-tema05\\barnehage\\kgdata.xlsx')
    sheet = workbook.active
    barnehager = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        barnehage = {
            'id': row[0],                        # ID
            'navn': row[1],                      # Navn
            'antall_plasser': row[2],            # Total # plasser
            'antall_ledige_plasser': row[3],     # tilgjengelige plasser
        }
        barnehager.append(barnehage)

    return barnehager
    
@app.route('/soeknader')
def soeknader():
    # Define the daycare centers with their available spots
    barnehager = [
        {'id': 1, 'navn': 'Sunshine Preschool', 'antall_plasser': 50, 'antall_ledige_plasser': 15},
        {'id': 2, 'navn': 'Happy Days Nursery', 'antall_plasser': 25, 'antall_ledige_plasser': 2},
        {'id': 3, 'navn': '123 Learning Center', 'antall_plasser': 35, 'antall_ledige_plasser': 4},
        {'id': 4, 'navn': 'ABC Kindergarten', 'antall_plasser': 12, 'antall_ledige_plasser': 0},
        {'id': 5, 'navn': 'Tiny Tots Academy', 'antall_plasser': 15, 'antall_ledige_plasser': 5},
        {'id': 6, 'navn': 'Giggles and Grins Childcare', 'antall_plasser': 10, 'antall_ledige_plasser': 0},
        {'id': 7, 'navn': 'Playful Pals Daycare', 'antall_plasser': 40, 'antall_ledige_plasser': 6},
    ]

    # Fetch applications, with each application specifying the desired daycare
    soeknader = hent_alle_soeknader()

    for soeknad in soeknader:
        barnehage_id = soeknad.get('barnehage_id')

        # Find the specific daycare the application is targeting
        valgt_barnehage = next((b for b in barnehager if b['id'] == barnehage_id), None)

        if valgt_barnehage and valgt_barnehage['antall_ledige_plasser'] > 0:
            # Assign "TILBUD" status and decrement available spots
            soeknad['status'] = "TILBUD"
            valgt_barnehage['antall_ledige_plasser'] -= 1
        else:
            # Assign "AVSLAG" if no spots are available
            soeknad['status'] = "AVSLAG"

    # Render the results in the soeknader.html template
    return render_template('soeknader.html', soeknader=soeknader)


@app.route('/svar')
def svar():
    information = session.get('information', {})
    barnehage_id = information.get('barnehage_id')  # Hent valgt barnehage fra søknadsinfo
    
    # Hardkodet liste over barnehager og ledige plasser
    barnehager = [
        {'id': 1, 'navn': 'Sunshine Preschool', 'antall_ledige_plasser': 15},
        {'id': 2, 'navn': 'Happy Days Nursery', 'antall_ledige_plasser': 2},
        {'id': 3, 'navn': '123 Learning Center', 'antall_ledige_plasser': 4},
        {'id': 4, 'navn': 'ABC Kindergarten', 'antall_ledige_plasser': 0},
        {'id': 5, 'navn': 'Tiny Tots Academy', 'antall_ledige_plasser': 5},
        {'id': 6, 'navn': 'Giggles and Grins Childcare', 'antall_ledige_plasser': 0},
        {'id': 7, 'navn': 'Playful Pals Daycare', 'antall_ledige_plasser': 6},
    ]

    # Finn barnehagen som matcher `barnehage_id`
    valgt_barnehage = next((b for b in barnehager if b['id'] == barnehage_id), None)

    # Bestem resultat basert på om det er ledige plasser
    if valgt_barnehage and valgt_barnehage['antall_ledige_plasser'] > 0:
        resultat = "TILBUD"
    else:
        resultat = "AVSLAG"

    return render_template('svar.html', resultat=resultat)



'''
@app.route('/svar')
def svar():
    information = session.get('information', {})
    resultat = "TILBUD" if antall_ledige_plasser > 0 else "AVSLAG"
    return render_template('svar.html', resultat=resultat)
    '''

@app.route('/statistikk', methods=['GET', 'POST'])
def statistikk():
    chart_json = None
    kommune = None

    if request.method == 'POST':
        kommune = request.form['kommune']
        print(f"Kommune valgt: {kommune}")  # Sjekk at kommunen hentes riktig
        filsti = r"C:\oblig5\is114-tema05\barnehage\ssb-barnehager-2015-2023-alder-1-2-aar.xlsm"
        
        try:
            # Les data fra Excel
            df = pd.read_excel(filsti, sheet_name='Sheet1')
            print("Excel-fil lastet inn vellykket")

            # Filtrer data for valgt kommune
            df_kommune = df[df['Kommune'] == kommune]
            if df_kommune.empty:
                print(f"Ingen data funnet for kommunen: {kommune}")
                return render_template('statistikk.html', chart_json=None, kommune=kommune, error=f"Ingen data funnet for {kommune}")

            # Konverter kolonneoverskriftene (år) til rader og sett opp data for Altair
            df_kommune = df_kommune.melt(id_vars=["Kommune"], var_name="År", value_name="Andel barn 1-2 år i barnehage")
            df_kommune = df_kommune.dropna()  # Fjern rader med NaN-verdier

            # Generer Altair-graf
            chart = alt.Chart(df_kommune).mark_line().encode(
                x='År:O',
                y='Andel barn 1-2 år i barnehage:Q',
                tooltip=['År', 'Andel barn 1-2 år i barnehage']
            ).properties(
                title=f"Utvikling i andel barn (1-2 år) i barnehage i {kommune} (2015-2023)"
            ).interactive()

            # Konverter graf til JSON for å sende til HTML
            chart_json = chart.to_json()
            print("Graf generert og konvertert til JSON vellykket")

        except Exception as e:
            print(f"En feil oppstod: {e}")
            return render_template('statistikk.html', chart_json=None, kommune=kommune, error=str(e))

    return render_template('statistikk.html', chart_json=chart_json, kommune=kommune)



@app.route('/commit')
def commit():
    commit_all()
    return render_template('commit.html')




"""
Referanser
[1] https://stackoverflow.com/questions/21668481/difference-between-render-template-and-redirect
"""

"""
Søkeuttrykk

"""